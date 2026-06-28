"""
ECP2.0 国网冀北电力物资采购数据采集流水线

流程:
  Phase 1: 从 noteList API 采集公告元数据 → bid_notices 表
  Phase 2: 关键词分类 (物资/服务/工程)
  Phase 3: 下载货物清单 ZIP → data/excels/{code}.zip
  Phase 4: 解压 → 识别货物清单XLSX → 解析 → bid_items 表
  Phase 5: 复核检查 → 输出差异报告

执行:
  python src/pipeline.py              # 增量模式 (只下载未处理的)
  python src/pipeline.py --full       # 全量模式 (采集全部)
  python src/pipeline.py --verify     # 仅复核检查
"""
import sys, os, io, re, json, time, zipfile, hashlib, subprocess, shutil, sqlite3
from datetime import datetime
from collections import Counter
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import requests
import openpyxl

from crawler.ecp_client import EcpClient, ORG_MAP
from db.schema import init_db, get_connection

# ============================================================
# 配置
# ============================================================
TARGET_ORG_ID = ORG_MAP["国网冀北电力有限公司"]
DOWNLOAD_URL = "https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/downLoadBid"
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "data", "excels")
EXTRACT_DIR = os.path.join(PROJECT_ROOT, "data", "extracted")
DB_PATH = os.path.join(PROJECT_ROOT, "data", "ecp_data.db")
FAILED_LOG = os.path.join(PROJECT_ROOT, "data", "failed_parse.json")

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
    "Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",
}

MATERIAL_KW = [
    "物资", "设备", "材料", "电缆", "变压器", "开关柜",
    "组合电器", "GIS", "断路器", "互感器", "避雷器",
    "绝缘子", "电容器", "电抗器", "导线", "光缆",
    "铁塔", "钢管杆", "金具", "线夹", "箱式变电站",
    "环网柜", "配电终端", "电能表",
]

# ============================================================
# Excel 列映射 — 基于表头自动检测, 而非硬编码列号
# ============================================================
# 货物清单表头签名 (两种格式)
GOODS_LIST_SIGNATURES = ["物资名称", "物料描述"]


def detect_column_map(headers: list[str]) -> dict[str, int] | None:
    """
    根据表头名称自动检测列索引。
    支持格式1: 23/27列招标公告货物清单 (物资名称)
    支持格式2: 零星物资/竞争性谈判货物清单 (物料描述/分标名称)
    返回 {字段名: 列索引}, 不是货物清单则返回 None。
    """
    has_signature = any(sig in headers for sig in GOODS_LIST_SIGNATURES)
    if not has_signature:
        return None

    col_map = {}
    for i, h in enumerate(headers):
        h = h.strip().replace(' ', '').replace('\n', '')
        if '分标编号' in h:
            col_map['sub_bid_code'] = i
        elif h == '包名称' or '分包名称' in h:
            col_map['package_name'] = i
        elif '分包编号' in h:
            col_map['package_no'] = i
        elif '项目单位' in h:
            col_map['project_org'] = i
        elif '需求单位' in h:
            col_map['demand_org'] = i
        elif '项目名称' in h:
            col_map['project_name'] = i
        elif '电压等级' in h:
            col_map['voltage_level'] = i
        elif h == '物资名称':
            col_map['material_name'] = i
        elif '物料描述' in h and col_map.get('material_name') is None:
            # 零星物资格式用"物料描述"代替"物资名称"
            col_map['material_name'] = i
        elif '物资描述' in h:
            col_map['material_desc'] = i
        elif h == '单位' or h == '计量单位':
            col_map['unit'] = i
        elif h == '数量':
            col_map['quantity'] = i
        elif '首批' in h:
            col_map['delivery_first'] = i
        elif '最后' in h:
            col_map['delivery_last'] = i
        elif '交货地点' in h:
            col_map['delivery_place'] = i
        elif '交货方式' in h:
            col_map['delivery_method'] = i
        elif h == '备注':
            col_map['remark'] = i
        elif '物料编码' in h or '细分物料编码' in h:
            col_map['material_code'] = i
        elif '扩展描述' in h:
            col_map['extended_desc'] = i
        elif '技术规范' in h:
            col_map['tech_spec_code'] = i
        elif '分标名称' in h and '分标编号' not in h:
            col_map['sub_bid_name_col'] = i  # Sheet名之外的分标名列

    # 至少需要物名称
    if 'material_name' not in col_map:
        return None
    return col_map


# 已知的硬编码修正 (notice_id → {sheet_name → {col_name → col_index}})
# 当自动检测失败时使用。从 failed_parse.json 分析后填入。
HARDCODED_FIXES: dict[str, dict] = {}

def load_hardcoded_fixes():
    """从配置文件加载硬编码修正"""
    global HARDCODED_FIXES
    fix_path = os.path.join(PROJECT_ROOT, "data", "hardcoded_fixes.json")
    if os.path.exists(fix_path):
        with open(fix_path, 'r', encoding='utf-8') as f:
            HARDCODED_FIXES = json.load(f)
    return HARDCODED_FIXES


def save_hardcoded_fixes():
    """保存硬编码修正"""
    fix_path = os.path.join(PROJECT_ROOT, "data", "hardcoded_fixes.json")
    with open(fix_path, 'w', encoding='utf-8') as f:
        json.dump(HARDCODED_FIXES, f, ensure_ascii=False, indent=2)


# ============================================================
# 工具函数
# ============================================================
def safe_cell(row: tuple, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    return str(val).strip() if val is not None else None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def extract_city(org_name: str) -> str | None:
    if not org_name:
        return None
    for city in ['唐山', '承德', '张家口', '秦皇岛', '廊坊', '北京']:
        if city in org_name:
            return city
    return None


# ============================================================
# Phase 1+2: 列表采集
# ============================================================
def phase1_collect(conn):
    print("\n" + "=" * 60)
    print("Phase 1+2: 公告列表采集 + 分类")
    print("=" * 60)

    client = EcpClient(timeout=30, retry=3)
    result = client.query_all(org_id=TARGET_ORG_ID, page_size=50)
    cursor = conn.cursor()
    material_count = 0

    for n in result.notices:
        cat = "material" if any(kw in n.title for kw in MATERIAL_KW) else "service"
        if "工程" in n.title and "服务" not in n.title:
            cat = "engineering"

        batch = ""
        m = re.search(r"第([一二三四五六七八九十\d]+)次", n.title)
        if m:
            batch = f"第{m.group(1)}次"

        year = None
        m = re.search(r"(20\d{2})", n.title)
        if m:
            year = int(m.group(1))

        # 使用 INSERT OR IGNORE + UPDATE 保留 excel_status
        cursor.execute("""
            INSERT OR IGNORE INTO bid_notices
                (notice_id, title, code, publish_org_name, org_id,
                 notice_publish_time, notice_type, notice_type_name,
                 doctype, doc_id, doc_url, zbflag,
                 category, bid_batch, bid_year, fetched_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (
            n.notice_id, n.title, n.code, n.publish_org_name, n.org_id,
            n.notice_publish_time, n.notice_type, n.notice_type_name,
            n.doctype, n.first_page_doc_id, n.doc_url, 0,
            cat, batch, year,
        ))
        # 更新元数据（但不覆盖 excel_status）
        cursor.execute("""
            UPDATE bid_notices
            SET title=?, code=?, publish_org_name=?,
                notice_publish_time=?, notice_type=?, notice_type_name=?,
                doctype=?, doc_id=?, doc_url=?,
                category=?, bid_batch=?, bid_year=?,
                updated_at=datetime('now')
            WHERE notice_id=?
        """, (
            n.title, n.code, n.publish_org_name,
            n.notice_publish_time, n.notice_type, n.notice_type_name,
            n.doctype, n.first_page_doc_id, n.doc_url,
            cat, batch, year,
            n.notice_id,
        ))
        if cat == "material":
            material_count += 1

    conn.commit()
    print(f"入库: {len(result.notices)} 条 (物资类: {material_count})")
    return material_count


# ============================================================
# Phase 3: 下载
# ============================================================
def phase3_download(conn, full_mode: bool = False) -> list[dict]:
    print("\n" + "=" * 60)
    print("Phase 3: 下载货物清单ZIP")
    print("=" * 60)

    cursor = conn.cursor()
    cursor.execute("""
        SELECT notice_id, title, code, notice_publish_time, doc_id
        FROM bid_notices
        WHERE category = 'material' AND doctype = 'doci-bid'
        ORDER BY notice_publish_time DESC
    """)
    notices = [dict(zip(['notice_id','title','code','pub_time','doc_id'], r))
               for r in cursor]
    print(f"物资正刊公告: {len(notices)} 条")

    os.makedirs(EXCEL_DIR, exist_ok=True)
    tasks = []
    new_downloads = 0

    for n in notices:
        notice_id = n['notice_id']
        code = n['code'] or notice_id

        # 文件名: code唯一则用code, 否则加noticeId后缀
        cursor.execute("SELECT COUNT(*) FROM bid_notices WHERE code=?", (code,))
        if cursor.fetchone()[0] > 1:
            zip_name = f"{code}_{notice_id}"
        else:
            zip_name = code

        zip_path = os.path.join(EXCEL_DIR, f"{zip_name}.zip")

        if os.path.exists(zip_path) and os.path.getsize(zip_path) > 1000:
            tasks.append({**n, 'zip_path': zip_path, 'zip_name': zip_name, 'status': 'exists'})
            # 确保已有文件的状态正确
            cursor.execute(
                "UPDATE bid_notices SET excel_status='downloaded' WHERE notice_id=? AND excel_status='pending'",
                (notice_id,))
            continue

        if full_mode or not os.path.exists(zip_path):
            url = f"{DOWNLOAD_URL}?noticeId={notice_id}&noticeDetId="
            try:
                resp = requests.get(url, headers=HTTP_HEADERS, timeout=30)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    with open(zip_path, 'wb') as f:
                        f.write(resp.content)
                    tasks.append({**n, 'zip_path': zip_path, 'zip_name': zip_name,
                                  'status': 'downloaded'})
                    cursor.execute(
                        "UPDATE bid_notices SET excel_status='downloaded' WHERE notice_id=?",
                        (notice_id,))
                    new_downloads += 1
                    time.sleep(0.3)
                else:
                    tasks.append({**n, 'zip_path': None, 'zip_name': zip_name,
                                  'status': f'empty({len(resp.content)}b)'})
                    cursor.execute(
                        "UPDATE bid_notices SET excel_status='no_file' WHERE notice_id=?",
                        (notice_id,))
            except Exception as e:
                tasks.append({**n, 'zip_path': None, 'zip_name': zip_name,
                              'status': f'error: {str(e)[:50]}'})
                cursor.execute(
                    "UPDATE bid_notices SET excel_status='no_file' WHERE notice_id=?",
                    (notice_id,))
        conn.commit()

    conn.commit()
    done = sum(1 for t in tasks if t['status'] in ('downloaded', 'exists'))
    skipped = sum(1 for t in tasks if t['zip_path'] is None)
    print(f"可处理: {done} 条 (新下载: {new_downloads})")
    print(f"跳过: {skipped} 条")
    return tasks


# ============================================================
# Phase 4: 解析
# ============================================================
def phase4_parse(conn, tasks: list[dict]) -> dict:
    print("\n" + "=" * 60)
    print("Phase 4: 解压 + 解析货物清单")
    print("=" * 60)

    cursor = conn.cursor()
    os.makedirs(EXTRACT_DIR, exist_ok=True)
    load_hardcoded_fixes()

    stats = {'success': 0, 'failed': 0, 'skipped': 0, 'total_items': 0}
    failed_entries = []
    all_materials = Counter()
    all_units = Counter()
    all_cities = Counter()

    for i, task in enumerate(tasks):
        notice_id = task['notice_id']
        title = task['title'][:70]

        if task['zip_path'] is None:
            stats['skipped'] += 1
            continue

        zip_size = os.path.getsize(task['zip_path']) / 1024
        if zip_size <= 0:
            stats['skipped'] += 1
            continue

        # 检查是否已解析 (增量模式跳过)
        cursor.execute("SELECT COUNT(*) FROM bid_items WHERE notice_id=?", (notice_id,))
        already_parsed = cursor.fetchone()[0]

        if already_parsed > 0:
            print(f"[{i+1}/{len(tasks)}] {task['pub_time']} {title[:55]}... "
                  f"({already_parsed}条已入库, 跳过)")
            stats['success'] += 1
            continue

        print(f"[{i+1}/{len(tasks)}] {task['pub_time']} {title[:55]}...")

        # 纯Python内存解压: 递归处理嵌套ZIP, 完全避免文件系统乱码问题
        goods_list_xlsx = _extract_goods_xlsx_memory(task['zip_path'])

        if not goods_list_xlsx:
            print(f"  -> 未找到货物清单XLSX")
            cursor.execute(
                "UPDATE bid_notices SET excel_status='no_file' WHERE notice_id=?",
                (notice_id,))
            conn.commit()
            stats['failed'] += 1
            failed_entries.append({'notice_id': notice_id, 'title': title,
                                    'reason': 'no_goods_list_xlsx'})
            continue

        # 保存Excel到 data/excels (goods_list_xlsx是临时文件)
        excel_name = task['zip_name'] + ".xlsx"
        saved_excel = os.path.join(EXCEL_DIR, excel_name)
        shutil.copy2(goods_list_xlsx, saved_excel)
        try:
            os.unlink(goods_list_xlsx)  # 删除临时文件
        except Exception:
            pass

        # 解析
        items, parse_errors = parse_goods_list(notice_id, saved_excel)

        if not items and not parse_errors:
            parse_errors.append("all_sheets_no_goods_list_signature")

        if not items:
            print(f"  -> 解析到0条 (原因: {parse_errors[:2]})")
            cursor.execute(
                "UPDATE bid_notices SET excel_status='parse_failed' WHERE notice_id=?",
                (notice_id,))
            conn.commit()
            stats['failed'] += 1
            failed_entries.append({
                'notice_id': notice_id, 'title': title,
                'reason': 'zero_items',
                'errors': parse_errors,
                'excel_path': saved_excel,
                'zip_name': task['zip_name'],
            })
            continue

        # 入库 bid_items
        inserted = 0
        for item in items:
            try:
                cursor.execute("""
                    INSERT OR IGNORE INTO bid_items
                        (notice_id, sub_bid_code, sub_bid_name, package_no,
                         material_name, material_desc, demand_quantity, unit,
                         project_org_name, delivery_place, remark,
                         material_code, extended_desc,
                         source, source_file)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    notice_id,
                    item.get('sub_bid_code'), item.get('sub_bid_name'),
                    item.get('package_no'), item.get('material_name'),
                    item.get('material_desc'), item.get('demand_quantity'),
                    item.get('unit'), item.get('project_org_name'),
                    item.get('delivery_place'), item.get('remark'),
                    item.get('material_code'), item.get('extended_desc'),
                    'goods_list_xlsx', excel_name,
                ))
                inserted += cursor.rowcount

                if item.get('material_name'):
                    short = item['material_name'].split(',')[0].split('(')[0].strip()
                    all_materials[short] += 1
                if item.get('unit'):
                    all_units[item['unit']] += 1
                city = extract_city(item.get('project_org_name', '') or '')
                if city:
                    all_cities[city] += 1
            except Exception as e:
                parse_errors.append(f"insert: {str(e)[:100]}")

        conn.commit()

        # 入库项目单位
        orgs_seen = set()
        for item in items:
            for key in ('project_org_name', 'demand_org_name'):
                name = item.get(key)
                if name and name not in orgs_seen:
                    orgs_seen.add(name)
                    city = extract_city(name)
                    cursor.execute("""
                        INSERT OR IGNORE INTO org_units
                            (parent_org_id, parent_org_name, org_name, city, province,
                             first_seen_notice_id, first_seen_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (TARGET_ORG_ID, "国网冀北电力有限公司", name,
                          city, "河北省" if city and city != "北京" else "北京市",
                          notice_id, task['pub_time']))
        conn.commit()

        # 标记解析成功
        cursor.execute("""
            UPDATE bid_notices
            SET excel_status='parsed', excel_path=?, detail_fetched=1,
                detail_fetched_at=datetime('now')
            WHERE notice_id=?
        """, (excel_name, notice_id))
        conn.commit()

        stats['success'] += 1
        stats['total_items'] += len(items)
        print(f"  -> {inserted}/{len(items)} 条, {len(orgs_seen)} 单位"
              f" (Excel: {excel_name})")
        if parse_errors:
            print(f"  -> ⚠️ {len(parse_errors)} 个警告: {parse_errors[:2]}")

    # 保存失败记录
    if failed_entries:
        with open(FAILED_LOG, 'w', encoding='utf-8') as f:
            json.dump(failed_entries, f, ensure_ascii=False, indent=2)
        print(f"\n失败记录: {FAILED_LOG} ({len(failed_entries)} 条)")

    stats['failed_entries'] = failed_entries
    stats['materials'] = all_materials
    stats['units'] = all_units
    stats['cities'] = all_cities
    return stats


def _extract_goods_xlsx_memory(zip_path: str) -> str | None:
    """
    纯Python内存解压ZIP, 递归处理嵌套ZIP, 找到货物清单XLSX。
    完全避免文件系统乱码文件名问题。返回临时XLSX路径或None。
    """
    import tempfile, zlib, struct

    def _raw_read(zf, info):
        """读取ZIP条目, 绕过文件名编码校验, 正确处理deflate"""
        with open(zf.filename, 'rb') as f:
            f.seek(info.header_offset + 26)
            name_len = struct.unpack('<H', f.read(2))[0]
            extra_len = struct.unpack('<H', f.read(2))[0]
            f.seek(info.header_offset + 30 + name_len + extra_len)
            compressed = f.read(info.compress_size)
        if info.compress_type == zipfile.ZIP_STORED:
            return compressed
        elif info.compress_type == zipfile.ZIP_DEFLATED:
            return zlib.decompress(compressed, -zlib.MAX_WBITS)
        return compressed

    def _search(zf, depth=0):
        if depth > 3:
            return None
        entries = list(zf.infolist())
        # 先找XLSX
        for info in entries:
            if info.filename.lower().endswith('.xlsx') and info.file_size > 500:
                try:
                    data = _raw_read(zf, info)
                    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                    if wb.sheetnames:
                        ws = wb[wb.sheetnames[0]]
                        h = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
                        if any(sig in h for sig in GOODS_LIST_SIGNATURES):
                            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                            tmp.write(data)
                            tmp.close()
                            wb.close()
                            return tmp.name
                    wb.close()
                except Exception:
                    continue
        # 再递归嵌套ZIP
        for info in entries:
            if info.filename.lower().endswith('.zip') and info.file_size > 100:
                try:
                    data = _raw_read(zf, info)
                    tmpz = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
                    tmpz.write(data)
                    tmpz.close()
                    with zipfile.ZipFile(tmpz.name, 'r') as izf:
                        result = _search(izf, depth + 1)
                    try: os.unlink(tmpz.name)
                    except: pass
                    if result:
                        return result
                except Exception:
                    continue
        return None

    with zipfile.ZipFile(zip_path, 'r') as zf:
        return _search(zf)


def _find_files(directory: str, suffix: str) -> list[str]:
    """递归查找指定后缀的文件, 使用os.scandir处理乱码文件名"""
    result = []
    try:
        with os.scandir(directory) as it:
            for entry in it:
                if entry.is_file() and entry.name.lower().endswith(suffix):
                    result.append(os.path.join(directory, entry.name))
                elif entry.is_dir():
                    result.extend(_find_files(entry.path, suffix))
    except (OSError, FileNotFoundError):
        pass
    return result


def _find_goods_xlsx_in_zip(zip_path: str) -> str | None:
    """
    纯Python内存解压ZIP, 递归处理嵌套ZIP, 找到货物清单XLSX。
    使用monkey-patch绕过zipfile的文件名编码校验。
    """
    import tempfile, zlib, struct

    # Monkey-patch: 绕过文件名编码校验
    _original_open = zipfile.ZipFile.open
    def _patched_open(self, name, mode='r', *args, **kwargs):
        if isinstance(name, zipfile.ZipInfo):
            info = name
        else:
            info = self.getinfo(name)
        # 直接用header_offset读取原始字节, 跳过文件名校验
        with open(self.filename, 'rb') as raw_f:
            raw_f.seek(info.header_offset)
            sig = raw_f.read(4)
            if sig != b'PK\x03\x04':
                raise zipfile.BadZipFile('Bad local header')
            raw_f.read(22)
            comp_size = struct.unpack('<I', raw_f.read(4))[0]
            uncomp_size = struct.unpack('<I', raw_f.read(4))[0]
            name_len = struct.unpack('<H', raw_f.read(2))[0]
            extra_len = struct.unpack('<H', raw_f.read(2))[0]
            raw_f.seek(name_len + extra_len, 1)
            raw = raw_f.read(comp_size)
        if comp_size == uncomp_size:
            return io.BytesIO(raw)
        return io.BytesIO(zlib.decompress(raw, -15))

    zipfile.ZipFile.open = _patched_open

    def _extract_and_find(fp: str, depth: int = 0) -> str | None:
        if depth > 3:
            return None
        try:
            with zipfile.ZipFile(fp, 'r') as zf:
                for info in list(zf.infolist()):
                    name_lower = info.filename.lower()
                    if name_lower.endswith('.xlsx') and info.file_size > 500:
                        try:
                            xlsx_data = zf.read(info)
                            wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=True)
                            if wb.sheetnames:
                                ws = wb[wb.sheetnames[0]]
                                h_row = [str(c.value or '') for c in
                                         next(ws.iter_rows(min_row=1, max_row=1))]
                                if any(sig in h_row for sig in GOODS_LIST_SIGNATURES):
                                    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                                    tmp.write(xlsx_data)
                                    tmp.close()
                                    wb.close()
                                    return tmp.name
                            wb.close()
                        except Exception:
                            continue
                for info in list(zf.infolist()):
                    name_lower = info.filename.lower()
                    if name_lower.endswith('.zip') and info.file_size > 100:
                        try:
                            inner_data = zf.read(info)
                            tmp_zip = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
                            tmp_zip.write(inner_data)
                            tmp_zip.close()
                            result = _extract_and_find(tmp_zip.name, depth + 1)
                            try: os.unlink(tmp_zip.name)
                            except: pass
                            if result:
                                return result
                        except Exception:
                            continue
        except Exception:
            pass
        return None

    result = _extract_and_find(zip_path)
    zipfile.ZipFile.open = _original_open  # restore
    return result


def _extract_nested_zips(directory: str):
    """递归解压目录中的嵌套ZIP文件 (处理ZIP内含ZIP的情况)
    使用cmd dir命令替代os.walk以避免乱码文件名问题"""
    extracted_any = True
    while extracted_any:
        extracted_any = False
        # 用dir /s /b找所有ZIP文件 (避免os.walk乱码)
        result = subprocess.run(
            ['cmd', '/c', 'dir', '/s', '/b', os.path.join(directory, '*.zip')],
            capture_output=True, text=True, timeout=15, encoding='gbk', errors='replace')
        zip_paths = [p.strip() for p in result.stdout.split('\n') if p.strip().lower().endswith('.zip')]
        for fp in zip_paths:
            try:
                if os.path.getsize(fp) < 100:
                    continue
                dest = fp.rsplit('.', 1)[0]
                os.makedirs(dest, exist_ok=True)
                subprocess.run(['unzip', '-o', '-UU', fp, '-d', dest],
                              capture_output=True, timeout=30)
                os.remove(fp)
                extracted_any = True
            except Exception:
                pass


def parse_goods_list(notice_id: str, xlsx_path: str) -> tuple[list[dict], list[str]]:
    """
    解析货物清单Excel。
    自动检测表头列映射 + 硬编码修正。
    """
    items = []
    errors = []

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        fixes = HARDCODED_FIXES.get(notice_id, {})

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            headers = [str(c or '') for c in rows[0]]
            col_map = detect_column_map(headers)
            if col_map is None:
                continue  # 不是货物清单Sheet

            # 应用硬编码修正
            if sheet_name in fixes:
                col_map.update(fixes[sheet_name])

            for row_idx, row in enumerate(rows[1:], 2):
                try:
                    mat_name = safe_cell(row, col_map.get('material_name'))
                    if not mat_name or mat_name == 'None':
                        continue

                    # 合并附加信息到remark
                    remark_parts = []
                    for key, label in [('voltage_level', '电压等级'),
                                       ('delivery_method', '交货方式'),
                                       ('delivery_first', '首批交货'),
                                       ('delivery_last', '最后交货'),
                                       ('project_name', '项目'),
                                       ('tech_spec_code', '技术规范')]:
                        v = safe_cell(row, col_map.get(key))
                        if v:
                            remark_parts.append(f"{label}:{v}")

                    items.append({
                        'sub_bid_code': safe_cell(row, col_map.get('sub_bid_code')),
                        'sub_bid_name': safe_cell(row, col_map.get('sub_bid_name_col')) or sheet_name,
                        'package_no': (safe_cell(row, col_map.get('package_no')) or
                                       safe_cell(row, col_map.get('package_name'))),
                        'material_name': mat_name,
                        'material_desc': safe_cell(row, col_map.get('material_desc')),
                        'demand_quantity': safe_float(safe_cell(row, col_map.get('quantity'))),
                        'unit': safe_cell(row, col_map.get('unit')),
                        'project_org_name': safe_cell(row, col_map.get('project_org')),
                        'demand_org_name': safe_cell(row, col_map.get('demand_org')),
                        'delivery_place': safe_cell(row, col_map.get('delivery_place')),
                        'remark': '; '.join(remark_parts) if remark_parts else None,
                        'material_code': safe_cell(row, col_map.get('material_code')),
                        'extended_desc': safe_cell(row, col_map.get('extended_desc')),
                    })
                except Exception as e:
                    errors.append(f"[{sheet_name}]R{row_idx}: {e}")

        wb.close()
    except Exception as e:
        errors.append(f"open: {e}")

    return items, errors


# ============================================================
# Phase 5: 复核检查
# ============================================================
def phase5_verify(conn):
    print("\n" + "=" * 60)
    print("Phase 5: 复核检查")
    print("=" * 60)

    cursor = conn.cursor()

    # 物资正刊总数
    cursor.execute("""
        SELECT COUNT(*) FROM bid_notices
        WHERE category='material' AND doctype='doci-bid'
    """)
    total_material = cursor.fetchone()[0]

    # 状态分布
    cursor.execute("""
        SELECT excel_status, COUNT(*) FROM bid_notices
        WHERE category='material' AND doctype='doci-bid'
        GROUP BY excel_status
    """)
    status_dist = {r[0]: r[1] for r in cursor}

    cursor.execute("""
        SELECT COUNT(DISTINCT notice_id) FROM bid_items
    """)
    with_items = cursor.fetchone()[0]

    print(f"\n物资正刊: {total_material} 条")
    print(f"状态分布: pending={status_dist.get('pending',0)} | downloaded={status_dist.get('downloaded',0)} | parsed={status_dist.get('parsed',0)} | no_file={status_dist.get('no_file',0)} | parse_failed={status_dist.get('parse_failed',0)}")
    print(f"有明细: {with_items} 条公告")

    # 未处理列表 (不是 parsed 状态的)
    cursor.execute("""
        SELECT notice_id, title, notice_publish_time, code, excel_status
        FROM bid_notices
        WHERE category='material' AND doctype='doci-bid'
          AND excel_status != 'parsed'
        ORDER BY notice_publish_time DESC
    """)
    pending = [dict(zip(['nid','title','pt','code','status'], r)) for r in cursor]
    if pending:
        print(f"\n未完成解析 ({len(pending)} 条):")
        for p in pending[:10]:
            print(f"  [{p['status']}] [{p['pt']}] {p['title'][:65]}")
        if len(pending) > 10:
            print(f"  ... 还有 {len(pending)-10} 条")

    # 重复检查
    cursor.execute("""
        SELECT COUNT(*) FROM (
            SELECT notice_id, sub_bid_code, sub_bid_name, package_no, material_name
            FROM bid_items
            GROUP BY 1,2,3,4,5 HAVING COUNT(*) > 1
        )
    """)
    dup_groups = cursor.fetchone()[0]
    print(f"\n{'✅ 无重复' if dup_groups == 0 else f'⚠️ {dup_groups} 组重复'}")

    # 统计
    cursor.execute("SELECT COUNT(*) FROM bid_items")
    total_items = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM org_units")
    org_count = cursor.fetchone()[0]

    # 城市维度
    cursor.execute("""
        SELECT o.city, COUNT(DISTINCT i.notice_id), COUNT(i.id)
        FROM bid_items i
        JOIN org_units o ON i.project_org_name = o.org_name
        WHERE o.city IS NOT NULL
        GROUP BY o.city ORDER BY COUNT(i.id) DESC
    """)
    city_stats = cursor.fetchall()

    print(f"\n总明细: {total_items} 条 | 单位: {org_count} 个")
    print(f"城市分布:")
    for city, nc, ic in city_stats:
        print(f"  {city}: {ic} 条, {nc} 公告")

    # 失败记录
    if os.path.exists(FAILED_LOG):
        with open(FAILED_LOG, 'r', encoding='utf-8') as f:
            failed = json.load(f)
        print(f"\n失败记录 ({len(failed)} 条):")
        for fe in failed:
            print(f"  [{fe['notice_id']}] {fe['reason']}: {fe.get('title','')[:60]}")

    return {
        'total_material': total_material,
        'parsed': status_dist.get('parsed', 0),
        'with_items': with_items, 'pending': len(pending),
        'total_items': total_items, 'org_count': org_count,
        'status_dist': status_dist,
        'failed': len(json.load(open(FAILED_LOG, encoding='utf-8')))
                  if os.path.exists(FAILED_LOG) else 0,
    }


# ============================================================
# 主入口
# ============================================================
def main():
    full_mode = '--full' in sys.argv
    verify_only = '--verify' in sys.argv

    print(f"ECP2.0 物资采购数据流水线")
    print(f"模式: {'全量' if full_mode else '增量'}"
          f"{' (仅复核)' if verify_only else ''}")
    print(f"开始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    conn = init_db(DB_PATH)

    if verify_only:
        phase5_verify(conn)
        conn.close()
        return

    # 确保db schema中有下载追踪字段
    try:
        conn.execute("ALTER TABLE bid_notices ADD COLUMN excel_path TEXT")
    except:
        pass

    phase1_collect(conn)
    tasks = phase3_download(conn, full_mode=full_mode)
    stats = phase4_parse(conn, tasks)
    result = phase5_verify(conn)

    print(f"\n{'='*60}")
    print("流水线完成, 更新所有产物...")
    print(f"{'='*60}")
    sd = result.get('status_dist', {})
    print(f"物资公告: {result['total_material']} | parsed={sd.get('parsed',0)} downloaded={sd.get('downloaded',0)} no_file={sd.get('no_file',0)} failed={sd.get('parse_failed',0)} pending={sd.get('pending',0)}")
    print(f"物资明细: {result['total_items']} | 单位: {result['org_count']}")

    # --- 自动更新所有产物 ---
    _update_all_outputs()
    conn.close()


def _update_all_outputs():
    """每次流水线运行后自动更新所有产物"""
    print()

    # 1. 物资需求统计 + 图表 (独立进程, 避免stdout冲突)
    try:
        subprocess.run([sys.executable, os.path.join(PROJECT_ROOT, 'src', 'demand_stats.py')],
                      timeout=120, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
        print("  [OK] material_demand_stats + 图表")
    except Exception as e:
        print(f"  [SKIP] demand_stats: {e}")

    # 2. 未处理公告Excel
    try:
        _export_unprocessed()
        print("  [OK] unprocessed_notices.xlsx")
    except Exception as e:
        print(f"  [SKIP] unprocessed: {e}")

    # 3. CSV导出
    try:
        import csv
        out = os.path.join(PROJECT_ROOT, 'outputs', 'material_demand_stats.csv')
        os.makedirs(os.path.dirname(out), exist_ok=True)
        conn3 = sqlite3.connect(DB_PATH)
        c3 = conn3.cursor()
        c3.execute('SELECT material_name, unit, demand_month, demand_quantity, notice_count FROM material_demand_stats ORDER BY demand_month, material_name')
        with open(out, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['material_name','unit','demand_month','demand_quantity','notice_count'])
            w.writerows(c3)
        conn3.close()
        print(f"  [OK] material_demand_stats.csv")
    except Exception as e:
        print(f"  [SKIP] CSV: {e}")

    print(f"\n数据库: {DB_PATH}")
    print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


def _export_unprocessed():
    """导出未处理公告到Excel"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''SELECT notice_id, title, code, notice_publish_time, doctype,
                        excel_status, category
                 FROM bid_notices WHERE excel_status != 'parsed'
                 ORDER BY CASE excel_status WHEN 'no_file' THEN 1 WHEN 'parse_failed' THEN 2
                 WHEN 'downloaded' THEN 3 ELSE 4 END, notice_publish_time DESC''')
    rows = c.fetchall()
    conn.close()

    if not rows:
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '未处理公告'
    headers = ['发布时间','公告类型','数据类别','标题','项目编号','文档类型','处理状态','状态码','未成功原因','页面链接']

    hdr_fill = PatternFill(start_color='009688',end_color='009688',fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        c2 = ws.cell(row=1, column=col, value=h)
        c2.fill = hdr_fill; c2.font = hdr_font
        c2.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)

    fills = {'no_file':'FFF3CD','downloaded':'D4EDDA','parse_failed':'F8D7DA','pending':'E2E3E5'}

    for ri, (nid,title,code,pt,dt,status,cat) in enumerate(rows, 2):
        menu = '2018032900295987' if any(k in title for k in ['竞争性谈判','零星物资','询价','单一来源']) else '2018032700291334'
        url = f'https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/{dt}/{nid}_{menu}'

        reasons = []
        if dt == 'doci-change': reasons.append('变更公告，无独立货物清单')
        elif dt == 'doc-spec': reasons.append('特殊文档，无货物清单')
        if cat in ('service','engineering'): reasons.append('服务/工程类，无物资数据')
        if status == 'no_file': reasons.append('ZIP中无货物清单XLSX')
        elif status == 'parse_failed': reasons.append('XLSX解析失败')
        elif status == 'downloaded': reasons.append('XLSX存在但自动解析失败')
        elif status == 'pending' and dt == 'doci-change': reasons.append('正刊已解析，无需重复')
        elif status == 'pending': reasons.append('非物资采购公告')

        vals = [pt, '采购公告' if menu=='2018032900295987' else '招标公告',
                {'material':'物资','service':'服务','engineering':'工程','other':'其他'}.get(cat,cat),
                title, code or '', dt,
                {'pending':'未下载','downloaded':'下载成功未解析','no_file':'无货物清单','parse_failed':'解析失败'}.get(status,status),
                status, '; '.join(reasons) if reasons else '待分析', url]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical='center', wrap_text=(ci in (4,9,10)))
            if status in fills:
                cell.fill = PatternFill(start_color=fills[status],end_color=fills[status],fill_type='solid')

    for i,w in enumerate([12,10,8,65,25,12,18,10,55,80],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    out = os.path.join(PROJECT_ROOT, 'data', 'unprocessed_notices.xlsx')
    wb.save(out)


if __name__ == "__main__":
    main()

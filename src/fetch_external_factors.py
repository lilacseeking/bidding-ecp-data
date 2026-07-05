"""
批量获取外部影响因子数据：STF + RMPF + IEDF (2020-05 ~ 2026-06)
写入 SQLite + 导出 Excel

执行: python src/fetch_external_factors.py
"""
import sys, os, io, sqlite3
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import numpy as np
import pandas as pd
import akshare as ak
from datetime import datetime

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DB_PATH = os.path.join(PROJECT_ROOT, "data", "ecp_data.db")
OUTPUT_XLSX = os.path.join(PROJECT_ROOT, "outputs", "外部影响因子数据.xlsx")

START = "202005"
END = "202606"


def all_months_range(start: str, end: str) -> list[str]:
    """生成完整自然月列表"""
    months = []
    y, m = int(start[:4]), int(start[4:])
    y_end, m_end = int(end[:4]), int(end[4:])
    while (y < y_end) or (y == y_end and m <= m_end):
        months.append("{:04d}{:02d}".format(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


# ============================================================
# STF: 季节时间因子 (纯计算)
# ============================================================
def build_stf(months: list[str]) -> pd.DataFrame:
    rows = []
    for mm in months:
        m = int(mm[4:])
        rows.append({
            "month": mm,
            "month_sin": round(np.sin(2 * np.pi * m / 12), 6),
            "month_cos": round(np.cos(2 * np.pi * m / 12), 6),
        })
    return pd.DataFrame(rows)


# ============================================================
# RMPF: 原材料价格因子 (SHFE 铜/铝期货月均价)
# ============================================================
def build_rmpf(months: list[str]) -> pd.DataFrame:
    print("  获取铜期货日数据...")
    cu = ak.futures_zh_daily_sina(symbol="CU0")
    cu["date"] = pd.to_datetime(cu["date"])
    cu["month"] = cu["date"].dt.strftime("%Y%m")
    cu_monthly = cu.groupby("month")["close"].mean().reset_index()
    cu_monthly.columns = ["month", "copper_price"]

    print("  获取铝期货日数据...")
    al = ak.futures_zh_daily_sina(symbol="AL0")
    al["date"] = pd.to_datetime(al["date"])
    al["month"] = al["date"].dt.strftime("%Y%m")
    al_monthly = al.groupby("month")["close"].mean().reset_index()
    al_monthly.columns = ["month", "aluminum_price"]

    df = cu_monthly.merge(al_monthly, on="month", how="outer")
    df = df[df["month"].between(START, END)].copy()
    df["copper_price"] = df["copper_price"].round(2)
    df["aluminum_price"] = df["aluminum_price"].round(2)

    # 补全缺失月份 (前向填充)
    full = pd.DataFrame({"month": months})
    df = full.merge(df, on="month", how="left")
    df["copper_price"] = df["copper_price"].ffill()
    df["aluminum_price"] = df["aluminum_price"].ffill()

    print("  铜期货: {} 个月覆盖, 铝期货: {} 个月覆盖".format(
        df["copper_price"].notna().sum(), df["aluminum_price"].notna().sum()))
    return df[["month", "copper_price", "aluminum_price"]]


# ============================================================
# IEDF: 工业用电需求因子 (全社会用电量)
# ============================================================
def build_iedf(months: list[str]) -> pd.DataFrame:
    print("  获取全社会用电量月度数据...")
    raw = ak.macro_china_society_electricity()
    raw["统计时间"] = raw["统计时间"].astype(str)

    # 转换月份格式: "2020.5" → "202005"
    def fmt_month(val):
        parts = val.split(".")
        if len(parts) == 2:
            return "{:04d}{:02d}".format(int(parts[0]), int(parts[1]))
        return val

    raw["month"] = raw["统计时间"].apply(fmt_month)
    raw = raw[raw["month"].between(START, END)].copy()

    # 第二产业用电量（工业用电）+ 同比增速
    df = raw[["month", "第二产业用电量", "第二产业用电量同比"]].copy()
    df.columns = ["month", "industrial_elec", "industrial_elec_yoy"]
    df["industrial_elec"] = df["industrial_elec"].astype(float) / 1e8  # 转为亿kWh
    df["industrial_elec"] = df["industrial_elec"].round(2)
    df["industrial_elec_yoy"] = df["industrial_elec_yoy"].astype(float)

    # 确保所有月份存在
    full = pd.DataFrame({"month": months})
    df = full.merge(df, on="month", how="left")

    print("  IEDF: {} 个月有数据, {} 个月缺失".format(
        df["industrial_elec"].notna().sum(), df["industrial_elec"].isna().sum()))
    return df[["month", "industrial_elec", "industrial_elec_yoy"]]


# ============================================================
# 写入 SQLite
# ============================================================
def write_db(df_all: pd.DataFrame):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS external_factors (
            month               TEXT PRIMARY KEY,
            month_sin           REAL,
            month_cos           REAL,
            copper_price        REAL,
            aluminum_price      REAL,
            industrial_elec     REAL,
            industrial_elec_yoy REAL,
            grid_invest         REAL,
            wind_capacity       REAL,
            solar_capacity      REAL,
            created_at          TEXT DEFAULT (datetime('now'))
        )
    """)

    # 使用 INSERT OR REPLACE 增量更新
    for _, row in df_all.iterrows():
        c.execute("""
            INSERT OR REPLACE INTO external_factors
                (month, month_sin, month_cos, copper_price, aluminum_price,
                 industrial_elec, industrial_elec_yoy)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            row["month"],
            row.get("month_sin"),
            row.get("month_cos"),
            row.get("copper_price"),
            row.get("aluminum_price"),
            row.get("industrial_elec"),
            row.get("industrial_elec_yoy"),
        ))

    conn.commit()

    c.execute("SELECT COUNT(*) FROM external_factors")
    cnt = c.fetchone()[0]
    print("\n数据库 external_factors: {} 条".format(cnt))

    # 打印预览
    c.execute("SELECT * FROM external_factors ORDER BY month LIMIT 5")
    cols = [d[0] for d in c.description]
    print("\n前5条预览:")
    for row in c:
        vals = ["{}={}".format(c, v) for c, v in zip(cols, row) if v is not None]
        print("  " + " | ".join(vals[:8]))

    conn.close()


# ============================================================
# 导出 Excel
# ============================================================
def export_xlsx(df_all: pd.DataFrame):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)

    wb = openpyxl.Workbook()

    # ---- Sheet1: 全量数据 ----
    ws1 = wb.active
    ws1.title = "外部影响因子"

    hdr_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    value_fill = PatternFill(start_color="d4e6f1", end_color="d4e6f1", fill_type="solid")
    none_fill = PatternFill(start_color="fdf2e9", end_color="fdf2e9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )

    headers = [
        ("月份", 14), ("STF_sin", 12), ("STF_cos", 12),
        ("铜期货月均价(元/吨)", 22), ("铝期货月均价(元/吨)", 22),
        ("第二产业用电量(亿kWh)", 24), ("工业用电同比增速(%)", 20),
        ("电网基建投资(亿元)", 20), ("风电新增装机(MW)", 18), ("光伏新增装机(MW)", 18),
    ]

    for col, (h, width) in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws1.column_dimensions[get_column_letter(col)].width = width

    color_cols = {
        "month": "#eaf2f8", "month_sin": "#f9ebea", "month_cos": "#f9ebea",
        "copper_price": "#e8f8f5", "aluminum_price": "#e8f8f5",
        "industrial_elec": "#fef9e7", "industrial_elec_yoy": "#fef9e7",
        "grid_invest": "#f4ecf7", "wind_capacity": "#eaf2f8", "solar_capacity": "#eaf2f8",
    }

    col_map = {
        "month": 1, "month_sin": 2, "month_cos": 3,
        "copper_price": 4, "aluminum_price": 5,
        "industrial_elec": 6, "industrial_elec_yoy": 7,
        "grid_invest": 8, "wind_capacity": 9, "solar_capacity": 10,
    }

    for ri, (_, row) in enumerate(df_all.iterrows(), 2):
        for col_name, col_idx in col_map.items():
            val = row.get(col_name)
            cell = ws1.cell(row=ri, column=col_idx, value=val if val is not None else None)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if col_name in color_cols:
                cell.fill = PatternFill(start_color=color_cols[col_name].lstrip("#"),
                                        end_color=color_cols[col_name].lstrip("#"),
                                        fill_type="solid")
            if val is None and col_name in ("grid_invest", "wind_capacity", "solar_capacity"):
                cell.fill = none_fill
                cell.value = "待采集"

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ---- Sheet2: 数据说明 ----
    ws2 = wb.create_sheet("数据说明")
    notes = [
        "外部影响因子数据说明",
        "",
        "数据范围: 2020-05 ~ 2026-06 (74个自然月)",
        "生成日期: " + datetime.now().strftime("%Y-%m-%d"),
        "",
        "=== 已完成 ===",
        "STF (季节时间因子): month_sin = sin(month×2π/12), month_cos = cos(month×2π/12)",
        "RMPF (原材料价格因子): 上海期货交易所(SHFE)铜(CU0)/铝(AL0)期货主力合约日收盘价月均值",
        "  数据来源: AKShare → futures_zh_daily_sina",
        "IEDF (工业用电需求因子): 全国第二产业用电量月度数据及同比增速",
        "  数据来源: AKShare → macro_china_society_electricity (原始数据来自国家能源局)",
        "",
        "=== 待采集 ===",
        "GIDF (电网投资驱动因子): 全国电网基建投资完成额月度数据",
        "  数据来源: 中电联《全国电力工业统计数据》月报, 需从新闻通稿手动整理",
        "NEIF (新能源并网容量因子): 风电/光伏新增装机容量",
        "  数据来源: 国家能源局季度可再生能源并网运行情况, 需从季报通稿手动整理",
        "",
        "=== 缺失值处理 ===",
        "RMPF: 期货交易日的缺失月份使用前向填充(ffill)",
        "IEDF: 2020年1-9月的部分月份在原始数据中缺失(API仅返回约66个月)",
        "GIDF/NEIF: 标记为'待采集', 数据需后续手动补充",
    ]
    for i, note in enumerate(notes, 1):
        cell = ws2.cell(row=i, column=1, value=note)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif note.startswith("==="):
            cell.font = Font(bold=True, size=11)

    ws2.column_dimensions["A"].width = 90

    wb.save(OUTPUT_XLSX)
    print("Excel: {} (2 Sheets)".format(OUTPUT_XLSX))


# ============================================================
# 主入口
# ============================================================
def main():
    print("外部影响因子数据获取")
    print("时间: {}".format(datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    print("范围: {} ~ {}".format(START, END))

    months = all_months_range(START, END)
    print("自然月: {} 个".format(len(months)))

    # STF
    print("\n[1/3] STF 季节时间因子 (计算)...")
    df_stf = build_stf(months)

    # RMPF
    print("\n[2/3] RMPF 原材料价格因子 (AKShare SHFE)...")
    df_rmpf = build_rmpf(months)

    # IEDF
    print("\n[3/3] IEDF 工业用电需求因子 (AKShare 用电量)...")
    df_iedf = build_iedf(months)

    # 合并
    df_all = df_stf.merge(df_rmpf, on="month", how="left") \
                   .merge(df_iedf, on="month", how="left")

    # 添加待采集列占位
    df_all["grid_invest"] = None
    df_all["wind_capacity"] = None
    df_all["solar_capacity"] = None

    # 有效性检查
    print("\n" + "=" * 50)
    print("数据完整性检查:")
    for col in ["month_sin", "month_cos", "copper_price", "aluminum_price",
                 "industrial_elec", "industrial_elec_yoy"]:
        valid = df_all[col].notna().sum()
        print("  {}: {}/{} ({:.0%})".format(col, valid, len(df_all), valid / len(df_all)))

    # 写入 SQLite
    print("\n写入 SQLite...")
    write_db(df_all)

    # 导出 Excel
    print("导出 Excel...")
    export_xlsx(df_all)

    print("\n完成: {}".format(datetime.now().strftime("%Y-%m-%d %H:%M:%S")))


if __name__ == "__main__":
    main()

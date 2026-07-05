"""
物资需求统计 + 多维度可视化仪表盘

执行: python src/demand_stats.py
生成图表:
  1. 年度招标物资金额/数量趋势
  2. 物资大类分布饼图
  3. 物资大类年度堆叠柱状图
  4. 月度公告数量趋势
  5. TOP15 物资需求量排名
  6. 月度需求热力图 (物资大类 × 月份)
"""
import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import sqlite3
from collections import Counter, defaultdict
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
from datetime import datetime

from db.schema import init_db

if 'pipeline' not in sys.modules:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "ecp_data.db")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs", "figures")

import matplotlib.font_manager as fm
for fname in ['Microsoft YaHei', 'SimHei', 'KaiTi']:
    try:
        font_path = fm.findfont(fm.FontProperties(family=fname), fallback_to_default=False)
        if font_path:
            plt.rcParams['font.sans-serif'] = [fname] + plt.rcParams['font.sans-serif']
            break
    except Exception:
        continue
plt.rcParams['axes.unicode_minus'] = False
fm._load_fontmanager(try_read_cache=False)

# 物资大类映射
CATEGORY_MAP = {
    '变压器': ['变压器', '变压器台成套', '调压器'],
    '开关柜/环网柜': ['开关柜', '环网柜', '环网箱', '箱式变电站', '低压屏柜', '配电箱', '端子箱', '空屏柜', '电能表屏'],
    '电缆/导线': ['电缆', '导线', '光缆', '布电线', '钢芯', '扩径导线', '软铜绞线',
                 '预制光缆', '通信电缆', '控制电缆', '导地线'],
    '避雷器/绝缘子': ['避雷器', '绝缘子', '穿墙套管', '交流支柱绝缘子', '拉紧绝缘子'],
    '断路器/组合电器': ['断路器', '组合电器', 'GIS', '隔离开关', '负荷开关', '熔断器', '高压熔断器'],
    '保护/监控/自动化': ['保护', '故障录波', '监控系统', '在线监测', '自动化', '时间同步',
                     '测控', '生产管理', '运维管理', '调度', '智能变电站', '压板', '数字化'],
    '通信设备': ['通信', '光端机', '交换机', '集线器', '综合接入', '电话及电视会议',
                '通信单元', '网络设备'],
    '电源/蓄电池': ['电源', '蓄电池', 'UPS', '充电', '直流', '逆变'],
    '杆塔/金具/铁附件': ['水泥杆', '锥形水泥杆', '钢管杆', '铁塔', '金具', '铁构件',
                       '铁附件', '地脚螺栓', '防鸟设备', '接地模块', '航空障碍灯'],
    '消防/安防': ['消防', '火灾报警', '防火', '灭火', '图像监视', '视频监视', '视频在线', '防山火'],
    '仪器仪表': ['电能表', '监测', '传感器', '检漏', '定位', '气象', '检测'],
    '其他': [],
}


def classify_material(name: str) -> str:
    for cat, keywords in CATEGORY_MAP.items():
        if cat == '其他':
            continue
        for kw in keywords:
            if kw in name:
                return cat
    return '其他'


def build_demand_stats(conn):
    c = conn.cursor()
    c.execute("DELETE FROM material_demand_item")
    c.execute("""INSERT OR REPLACE INTO material_demand_item
        (material_name, unit, demand_month, demand_quantity, notice_count)
        SELECT CASE WHEN i.material_name LIKE '%,%'
                 THEN substr(i.material_name, 1, instr(i.material_name, ',') - 1)
                 ELSE i.material_name END,
            i.unit, i.demand_month, SUM(i.demand_quantity), COUNT(DISTINCT i.notice_id)
        FROM bid_items i
        WHERE i.demand_quantity IS NOT NULL AND i.unit IS NOT NULL AND i.material_name IS NOT NULL
        GROUP BY 1, 2, 3""")
    conn.commit()
    c.execute("DELETE FROM material_demand_total")
    c.execute("""INSERT OR REPLACE INTO material_demand_total
        (material_name, demand_month, demand_quantity, notice_count)
        SELECT material_name, demand_month, SUM(demand_quantity), MAX(notice_count)
        FROM material_demand_item GROUP BY material_name, demand_month""")
    conn.commit()
    c.execute("SELECT COUNT(*) FROM material_demand_item")
    print(f"material_demand_item: {c.fetchone()[0]} 条")


# ============================================================
# 图表1: 年度招标物资总量趋势 (按大类分色堆叠柱状图)
# ============================================================
def plot_yearly_category_stacked(conn):
    c = conn.cursor()
    c.execute("""SELECT demand_month, material_name, SUM(demand_quantity)
        FROM material_demand_total GROUP BY 1, 2 ORDER BY 1""")
    rows = c.fetchall()
    if not rows:
        return

    yearly_cat = defaultdict(lambda: defaultdict(float))
    for month, mat, qty in rows:
        year = month[:4]
        cat = classify_material(mat)
        yearly_cat[year][cat] += qty

    years = sorted(yearly_cat.keys())
    categories = ['电缆/导线', '杆塔/金具/铁附件', '开关柜/环网柜', '避雷器/绝缘子',
                  '保护/监控/自动化', '断路器/组合电器', '通信设备', '变压器',
                  '电源/蓄电池', '消防/安防', '仪器仪表', '其他']
    colors = ['#E53935', '#1E88E5', '#43A047', '#FB8C00', '#8E24AA',
              '#00ACC1', '#FDD835', '#6D4C41', '#546E7A', '#D81B60', '#3949AB', '#BDBDBD']

    data = []
    labels = []
    for cat in categories:
        row = [yearly_cat[y].get(cat, 0) for y in years]
        if any(v > 0 for v in row):
            data.append(row)
            labels.append(cat)

    if not data:
        return

    fig, ax = plt.subplots(figsize=(16, 7))
    x = range(len(years))
    bottom = np.zeros(len(years))
    for i, (cat_data, cat_label) in enumerate(zip(data, labels)):
        bars = ax.bar(x, cat_data, bottom=bottom, label=cat_label,
                     color=colors[i % len(colors)], width=0.6, edgecolor='white', linewidth=0.5)
        bottom += np.array(cat_data)
        for bar, val in zip(bars, cat_data):
            if val > max(cat_data) * 0.1:
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_y() + bar.get_height()/2,
                       f'{val/10000:.0f}万' if val >= 10000 else f'{val:.0f}',
                       ha='center', va='center', fontsize=7, fontweight='bold')

    ax.set_xticks(x)
    ax.set_xticklabels(years, fontsize=12)
    ax.set_title('国网冀北电力 — 年度招标物资总量趋势 (按大类堆叠)', fontsize=16, fontweight='bold', pad=15)
    ax.set_ylabel('需求数量 (台/套/千米等)', fontsize=12)
    ax.legend(loc='upper left', fontsize=9, ncol=3, framealpha=0.9)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    ax.set_ylim(0, bottom.max() * 1.15)

    path = os.path.join(OUTPUT_DIR, '年度招标物资总量趋势.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 图表2: 物资大类分布饼图
# ============================================================
def plot_category_pie(conn):
    c = conn.cursor()
    c.execute("SELECT material_name, SUM(demand_quantity) FROM material_demand_total GROUP BY 1")
    rows = c.fetchall()
    if not rows:
        return

    cat_totals = defaultdict(float)
    for mat, qty in rows:
        cat_totals[classify_material(mat)] += qty

    sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
    labels = [c for c, _ in sorted_cats]
    values = [v for _, v in sorted_cats]
    colors = ['#E53935', '#1E88E5', '#43A047', '#FB8C00', '#8E24AA',
              '#00ACC1', '#FDD835', '#6D4C41', '#546E7A', '#D81B60', '#3949AB', '#BDBDBD']

    fig, ax = plt.subplots(figsize=(10, 10))
    wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%',
        colors=colors[:len(labels)], startangle=90, pctdistance=0.85,
        wedgeprops=dict(width=0.4, edgecolor='white'))
    for t in autotexts:
        t.set_fontsize(9)
        t.set_fontweight('bold')
    for t in texts:
        t.set_fontsize(11)
    ax.set_title('国网冀北电力 — 物资大类需求分布 (2020-2026累计)', fontsize=16, fontweight='bold', pad=20)

    path = os.path.join(OUTPUT_DIR, '物资大类分布饼图.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 图表3: 月度公告数量趋势
# ============================================================
def plot_monthly_notice_count(conn):
    c = conn.cursor()
    c.execute("""SELECT demand_month, COUNT(DISTINCT notice_id) as cnt
        FROM bid_items GROUP BY demand_month ORDER BY demand_month""")
    rows = c.fetchall()
    if not rows:
        return

    months = [r[0] for r in rows]
    counts = [r[1] for r in rows]
    years = sorted(set(m[:4] for m in months))
    colors = plt.cm.tab20(np.linspace(0, 1, len(years)))

    fig, ax = plt.subplots(figsize=(18, 6))
    bars = ax.bar(range(len(months)), counts, width=0.8,
                  color=[colors[years.index(m[:4])] for m in months],
                  edgecolor='white', linewidth=0.3)

    # 年度分割线
    prev_year = months[0][:4]
    for i, m in enumerate(months):
        if m[:4] != prev_year:
            ax.axvline(i - 0.5, color='#999', linestyle='--', linewidth=0.8, alpha=0.7)
            prev_year = m[:4]

    tick_pos = list(range(0, len(months), max(1, len(months)//12)))
    tick_lbl = [f"{months[i][:4]}-{months[i][4:]}" for i in tick_pos]
    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_lbl, fontsize=8, rotation=45, ha='right')
    ax.set_title('国网冀北电力 — 月度招标物资公告数量趋势', fontsize=16, fontweight='bold')
    ax.set_ylabel('涉及公告数', fontsize=12)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 图例
    from matplotlib.patches import Patch
    legend_elements = [Patch(facecolor=colors[i], label=y) for i, y in enumerate(years)]
    ax.legend(handles=legend_elements, fontsize=9, loc='upper left', ncol=len(years))

    path = os.path.join(OUTPUT_DIR, '月度公告数量趋势.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 图表4: TOP15 物资需求量排名
# ============================================================
def plot_top15_materials(conn):
    c = conn.cursor()
    c.execute("""SELECT material_name, SUM(demand_quantity) as total
        FROM material_demand_total GROUP BY material_name
        ORDER BY total DESC LIMIT 15""")
    rows = c.fetchall()
    if not rows:
        return

    names = [r[0][:30] for r in reversed(rows)]
    values = [r[1] for r in reversed(rows)]
    cat_colors = []
    for r in reversed(rows):
        cat = classify_material(r[0])
        cat_colors.append({'电缆/导线': '#E53935', '杆塔/金具/铁附件': '#1E88E5',
                          '开关柜/环网柜': '#43A047', '避雷器/绝缘子': '#FB8C00',
                          '保护/监控/自动化': '#8E24AA', '断路器/组合电器': '#00ACC1',
                          '通信设备': '#FDD835', '变压器': '#6D4C41',
                          '电源/蓄电池': '#546E7A', '消防/安防': '#D81B60',
                          '仪器仪表': '#3949AB', '其他': '#BDBDBD'}.get(cat, '#999'))

    fig, ax = plt.subplots(figsize=(14, 8))
    bars = ax.barh(range(len(names)), values, color=cat_colors, edgecolor='white', height=0.7)
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, fontsize=10)
    ax.set_xlabel('需求数量', fontsize=12)
    ax.set_title('国网冀北电力 — TOP15 物资需求排名', fontsize=16, fontweight='bold')
    ax.grid(axis='x', alpha=0.3, linestyle='--')

    for bar, val in zip(bars, values):
        label = f'{val/10000:.1f}万' if val >= 10000 else f'{val:,.0f}'
        ax.text(bar.get_width() + max(values)*0.01, bar.get_y() + bar.get_height()/2,
               label, va='center', fontsize=9, fontweight='bold')

    # 图例
    from matplotlib.patches import Patch
    color_dict = {'电缆/导线': '#E53935', '杆塔/金具/铁附件': '#1E88E5',
                  '开关柜/环网柜': '#43A047', '避雷器/绝缘子': '#FB8C00',
                  '保护/监控/自动化': '#8E24AA', '断路器/组合电器': '#00ACC1',
                  '通信设备': '#FDD835', '变压器': '#6D4C41',
                  '电源/蓄电池': '#546E7A', '消防/安防': '#D81B60',
                  '仪器仪表': '#3949AB', '其他': '#BDBDBD'}
    unique_cats = list(set((classify_material(r[0]) for r in rows)))
    legend_elements = [Patch(facecolor=color_dict[cat], label=cat) for cat in unique_cats]
    ax.legend(handles=legend_elements, fontsize=9, loc='lower right')

    path = os.path.join(OUTPUT_DIR, 'TOP15物资需求量排名.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 图表5: 热度图 — 物资大类 × 年份需求热力
# ============================================================
def plot_category_heatmap(conn):
    c = conn.cursor()
    c.execute("SELECT demand_month, material_name, SUM(demand_quantity) FROM material_demand_total GROUP BY 1,2")
    rows = c.fetchall()
    if not rows:
        return

    yearly_cat = defaultdict(lambda: defaultdict(float))
    for month, mat, qty in rows:
        cat = classify_material(mat)
        yearly_cat[cat][month[:4]] += qty

    categories = [c for c in CATEGORY_MAP.keys() if c in yearly_cat]
    years = sorted(set(y for cat_data in yearly_cat.values() for y in cat_data))

    if not years or not categories:
        return

    data = np.zeros((len(categories), len(years)))
    for i, cat in enumerate(categories):
        for j, year in enumerate(years):
            data[i, j] = yearly_cat[cat].get(year, 0)

    # Normalize per row for better visualization
    data_log = np.log1p(data)

    fig, ax = plt.subplots(figsize=(12, 8))
    im = ax.imshow(data_log, cmap='YlOrRd', aspect='auto')

    ax.set_xticks(range(len(years)))
    ax.set_xticklabels(years, fontsize=11)
    ax.set_yticks(range(len(categories)))
    ax.set_yticklabels(categories, fontsize=11)

    # Annotate with actual values
    for i in range(len(categories)):
        for j in range(len(years)):
            val = data[i, j]
            if val > 0:
                label = f'{val/10000:.0f}万' if val >= 10000 else f'{val:.0f}'
                ax.text(j, i, label, ha='center', va='center', fontsize=8,
                       fontweight='bold', color='white' if data_log[i,j] > data_log.max()*0.6 else 'black')

    ax.set_title('国网冀北电力 — 物资大类 × 年度需求热力图', fontsize=16, fontweight='bold', pad=15)
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('log(1+需求量)', fontsize=10)

    path = os.path.join(OUTPUT_DIR, '物资需求年度热力图.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 图表6: TOP20 高频采购物资月度趋势 (完整自然月X轴, 0值填充)
# ============================================================
def plot_core_materials_trend(conn):
    c = conn.cursor()
    # 获取完整自然月范围
    c.execute("SELECT MIN(demand_month), MAX(demand_month) FROM bid_items WHERE demand_month IS NOT NULL")
    min_m, max_m = c.fetchone()
    all_months = []
    y, m = int(min_m[:4]), int(min_m[4:])
    y_end, m_end = int(max_m[:4]), int(max_m[4:])
    while (y < y_end) or (y == y_end and m <= m_end):
        all_months.append("{:04d}{:02d}".format(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    n_months = len(all_months)

    # TOP20 按总需求量排序
    c.execute("""SELECT material_name, SUM(demand_quantity) as total
        FROM material_demand_total GROUP BY material_name
        ORDER BY total DESC LIMIT 20""")
    top20 = c.fetchall()
    if not top20:
        return

    n_cols = 4
    n_rows = 5
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(36, 28))
    fig.suptitle('国网冀北电力 — TOP20 物资月度需求趋势 (完整自然月)', fontsize=18, fontweight='bold', y=0.99)

    # 20种颜色: Tableau 20 调色板
    colors = [
        '#E53935', '#1E88E5', '#43A047', '#FB8C00',
        '#8E24AA', '#00ACC1', '#F4511E', '#3949AB',
        '#7CB342', '#C0CA33', '#5E35B1', '#00897B',
        '#D81B60', '#6D4C41', '#546E7A', '#FDD835',
        '#039BE5', '#33A860', '#EF6C00', '#8D6E63',
    ]

    # 预查询所有物资的月度数据
    placeholders = ','.join(['?'] * 20)
    c.execute("""SELECT material_name, demand_month, demand_quantity
        FROM material_demand_total
        WHERE material_name IN ({})
        ORDER BY material_name, demand_month""".format(placeholders),
        [t[0] for t in top20])
    all_rows = c.fetchall()

    mat_month_data = {}
    for name, dm, qty in all_rows:
        if name not in mat_month_data:
            mat_month_data[name] = {}
        mat_month_data[name][dm] = qty

    for idx in range(20):
        ax = axes[idx // n_cols][idx % n_cols]
        mat_name = top20[idx][0]
        total = top20[idx][1]

        month_qty = mat_month_data.get(mat_name, {})
        quantities = [month_qty.get(m, 0) for m in all_months]
        nonzero_cnt = sum(1 for v in quantities if v > 0)

        ax.fill_between(range(n_months), quantities, alpha=0.25, color=colors[idx])
        ax.plot(range(n_months), quantities, 'o-', color=colors[idx],
                linewidth=1.5, markersize=2, markerfacecolor='white',
                markeredgewidth=0.5)

        # 标注峰值
        nonzero = [(i, v) for i, v in enumerate(quantities) if v > 0]
        if nonzero:
            mx_idx, mx_val = max(nonzero, key=lambda x: x[1])
            ax.annotate('{:.1f}万'.format(mx_val / 10000) if mx_val >= 10000 else '{:.0f}'.format(mx_val),
                       xy=(mx_idx, mx_val),
                       xytext=(0, 10), textcoords='offset points', fontsize=8,
                       ha='center', color=colors[idx], fontweight='bold')

        # X轴：每年标一个刻度
        year_starts = [i for i, m in enumerate(all_months) if m[4:] == '01']
        if idx >= (n_rows - 1) * n_cols:  # 最后一行显示年份标签
            year_labels = [all_months[i][:4] for i in year_starts]
            ax.set_xticks(year_starts)
            ax.set_xticklabels(year_labels, fontsize=7)
        else:
            ax.set_xticks(year_starts)
            ax.set_xticklabels([])
        for ys in year_starts:
            ax.axvline(ys, color='#ccc', linestyle='--', linewidth=0.4, alpha=0.4)

        # 标题: rank + 物资名 + 有数据月数 + 总量
        short_name = mat_name[:28] + '..' if len(mat_name) > 30 else mat_name
        ax.set_title('#{} {} | {}月 | {:.0f}万'.format(
            idx + 1, short_name, nonzero_cnt, total / 10000),
            fontsize=9, loc='left')
        ax.grid(axis='y', alpha=0.2, linestyle='--')
        ax.set_xlim(-0.5, n_months - 0.5)

        ymax = max(quantities) if max(quantities) > 0 else 1
        ax.set_ylim(-ymax * 0.02, ymax * 1.15)
        ax.tick_params(axis='y', labelsize=7)

    fig.tight_layout(rect=[0.02, 0.02, 0.98, 0.96])

    path = os.path.join(OUTPUT_DIR, '物资需求Top5月度趋势.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print('  [OK] {} ({} 个自然月, TOP20物资, 含0值)'.format(os.path.basename(path), n_months))


# ============================================================
# 图表7: 招标公告 vs 采购公告年度对比
# ============================================================
def plot_notice_type_comparison(conn):
    c = conn.cursor()
    c.execute("""SELECT notice_publish_time, title FROM bid_notices
        WHERE category='material' AND doctype='doci-bid'
        AND title NOT LIKE '%变更%' AND source_file IS NOT NULL""")
    rows = c.fetchall()

    monthly = defaultdict(lambda: {'招标公告': 0, '采购公告': 0})
    for pt, title in rows:
        month = pt[:4] + pt[5:7]
        tp = '采购公告' if any(k in (title or '') for k in ['竞争性谈判','零星物资','询价']) else '招标公告'
        monthly[month][tp] += 1

    months_sorted = sorted(monthly.keys())
    zb = [monthly[m]['招标公告'] for m in months_sorted]
    cg = [monthly[m]['采购公告'] for m in months_sorted]

    fig, ax = plt.subplots(figsize=(18, 6))
    x = range(len(months_sorted))
    ax.bar(x, zb, width=0.6, label='招标公告', color='#1E88E5', edgecolor='white')
    ax.bar(x, cg, width=0.6, bottom=zb, label='采购公告(竞争性谈判/零星物资)',
           color='#FF9800', edgecolor='white')

    tick_pos = list(range(0, len(months_sorted), max(1, len(months_sorted)//12)))
    tick_lbl = [f"{months_sorted[i][:4]}-{months_sorted[i][4:]}" for i in tick_pos]
    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_lbl, fontsize=8, rotation=45, ha='right')
    ax.set_title('国网冀北电力 — 招标公告 vs 采购公告月度分布', fontsize=16, fontweight='bold')
    ax.set_ylabel('公告数量', fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    path = os.path.join(OUTPUT_DIR, '公告类型月度对比.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  [OK] {os.path.basename(path)}')


# ============================================================
# 主入口
# ============================================================
def main():
    print("物资需求统计 + 多维度可视化")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    init_db(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    build_demand_stats(conn)

    # 生成所有图表
    print("\n生成图表:")
    plot_yearly_category_stacked(conn)
    plot_category_pie(conn)
    plot_monthly_notice_count(conn)
    plot_top15_materials(conn)
    plot_category_heatmap(conn)
    plot_core_materials_trend(conn)
    plot_notice_type_comparison(conn)

    # 导出Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "物资需求统计.xlsx")
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    c = conn.cursor()
    c.execute("""SELECT material_name, unit, demand_month, demand_quantity, notice_count
        FROM material_demand_item ORDER BY demand_month, material_name""")
    rows = c.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '物资需求统计'
    hdr_f = PatternFill(start_color='009688', end_color='009688', fill_type='solid')
    for col, h in enumerate(['物资名称','单位','月份','需求量','公告数'], 1):
        c2 = ws.cell(row=1, column=col, value=h)
        c2.fill = hdr_f; c2.font = Font(color='FFFFFF', bold=True)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1): ws.cell(row=ri, column=ci, value=val)
    ws.column_dimensions['A'].width = 50
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"\nExcel: {xlsx_path} ({len(rows)} 行)")

    conn.close()
    print(f"\n完成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()

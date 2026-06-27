"""
完整数据采集流水线
Phase 1: 列表采集 (noteList API, 无需登录) ✅
Phase 2: 分类筛选 ✅
Phase 3: 下载货物清单ZIP (downLoadBid API, GET请求, 无需登录!) ✅ 新发现
Phase 4: 解压ZIP + 解析Excel + 入库SQLite

执行: python src/pipeline_full.py
"""
import sys, os, io, re, json, time, zipfile, struct, tempfile, shutil
from datetime import datetime
from collections import Counter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import requests
import openpyxl

from crawler.ecp_client import EcpClient, ORG_MAP
from db.schema import init_db, get_connection

# ============================================================
# 配置
# ============================================================
TARGET_ORG_ID = ORG_MAP["国网冀北电力有限公司"]
ECP_DOWNLOAD = "https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/downLoadBid"
DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "downloads")
EXTRACT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "extracted")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
    "Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",
}

MATERIAL_KW = [
    "物资", "设备", "材料", "电缆", "变压器", "开关柜",
    "组合电器", "GIS", "断路器", "互感器", "避雷器",
    "绝缘子", "电容器", "电抗器", "导线", "光缆",
    "铁塔", "钢管杆", "金具", "线夹", "箱式变电站",
    "环网柜", "配电终端", "电能表",
]


# ============================================================
# Phase 1+2: 列表采集 (复用已有逻辑)
# ============================================================
def phase1_collect_list(conn):
    """采集冀北招标公告列表并分类"""
    print("\n" + "=" * 70)
    print("Phase 1+2: 采集公告列表 + 分类")
    print("=" * 70)

    client = EcpClient(timeout=30, retry=3)
    result = client.query_all(org_id=TARGET_ORG_ID, page_size=50)

    cursor = conn.cursor()
    material_ids = []
    stored = 0

    for n in result.notices:
        cat = "material" if any(kw in n.title for kw in MATERIAL_KW) else "service"
        if "工程" in n.title and "服务" not in n.title:
            cat = "engineering"

        batch = ""
        m = re.search(r"第([一二三四五六七八九十\d]+)次", n.title)
        if m:
            batch = f"第{m.group(1)}次"

        year = None
        m = re.search(r"(20\d{2})", n.title)
        if m:
            year = int(m.group(1))

        cursor.execute("""
            INSERT OR REPLACE INTO bid_notices
                (notice_id, title, code, publish_org_name, org_id,
                 notice_publish_time, notice_type, notice_type_name,
                 doctype, doc_id, doc_url, zbflag,
                 category, bid_batch, bid_year, fetched_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (
            n.notice_id, n.title, n.code, n.publish_org_name, n.org_id,
            n.notice_publish_time, n.notice_type, n.notice_type_name,
            n.doctype, n.first_page_doc_id, n.doc_url,
            0, cat, batch, year,
        ))
        stored += 1
        if cat == "material":
            material_ids.append((n.notice_id, n.title, n.notice_publish_time, n.code))

    conn.commit()
    print(f"存储: {stored} 条, 物资类: {len(material_ids)} 条")
    return material_ids


# ============================================================
# Phase 3: 下载货物清单ZIP
# ============================================================
def download_zip(notice_id: str) -> str | None:
    """下载公告的货物清单ZIP文件"""
    url = f"{ECP_DOWNLOAD}?noticeId={notice_id}&noticeDetId="

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    out_path = os.path.join(DOWNLOAD_DIR, f"{notice_id}.zip")

    # 跳过已下载的
    if os.path.exists(out_path) and os.path.getsize(out_path) > 1000:
        return out_path

    try:
        resp = requests.get(url, headers=HEADERS, timeout=30, stream=True)
        if resp.status_code == 200 and len(resp.content) > 1000:
            with open(out_path, 'wb') as f:
                f.write(resp.content)
            time.sleep(0.3)
            return out_path
    except Exception as e:
        pass
    return None


# ============================================================
# Phase 4: 解压ZIP + 解析Excel + 入库
# ============================================================
def extract_zip(zip_path: str) -> tuple[str | None, str | None]:
    """解压ZIP，返回(xlsx_path, doc_path)"""
    notice_id = os.path.basename(zip_path).replace('.zip', '')
    extract_subdir = os.path.join(EXTRACT_DIR, notice_id)
    os.makedirs(extract_subdir, exist_ok=True)

    xlsx_path = None
    doc_path = None

    try:
        # 使用unzip -UU忽略Unicode冲突
        import subprocess
        subprocess.run([
            'unzip', '-o', '-UU', zip_path, '-d', extract_subdir
        ], capture_output=True, timeout=30)

        for f in os.listdir(extract_subdir):
            full = os.path.join(extract_subdir, f)
            size = os.path.getsize(full)
            if f.endswith('.xlsx') and size > 50000:
                xlsx_path = full
            elif (f.endswith('.doc') or f.endswith('.docx')) and size > 50000:
                doc_path = full
    except Exception as e:
        pass

    return xlsx_path, doc_path


def parse_xlsx(xlsx_path: str) -> list[dict]:
    """解析货物清单Excel，返回物资明细列表"""
    items = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            if not rows:
                continue

            for row in rows:
                if not row or len(row) < 11:
                    continue

                material_name = str(row[7]).strip() if row[7] else None
                if not material_name or material_name == 'None':
                    continue

                item = {
                    'sub_bid_code': str(row[0]).strip() if row[0] else None,
                    'package_no': str(row[2]).strip() if row[2] else (
                        str(row[1]).strip() if row[1] else None
                    ),
                    'project_org_name': str(row[3]).strip() if row[3] else None,
                    'demand_org_name': str(row[4]).strip() if row[4] else None,
                    'project_name': str(row[5]).strip() if row[5] else None,
                    'voltage_level': str(row[6]).strip() if row[6] else None,
                    'material_name': material_name,
                    'material_desc': str(row[8]).strip() if row[8] else None,
                    'unit': str(row[9]).strip() if row[9] else None,
                    'demand_quantity': _safe_float(row[10]),
                    'delivery_date_first': str(row[11]).strip() if len(row) > 11 and row[11] else None,
                    'delivery_date_last': str(row[12]).strip() if len(row) > 12 and row[12] else None,
                    'delivery_place': str(row[13]).strip() if len(row) > 13 and row[13] else None,
                    'delivery_method': str(row[14]).strip() if len(row) > 14 and row[14] else None,
                    'sub_bid_name': sheet_name,
                    'source': 'goods_list_xlsx',
                    'source_file': os.path.basename(xlsx_path),
                }
                items.append(item)
        wb.close()
    except Exception as e:
        pass

    return items


def _safe_float(val) -> float | None:
    """安全转换为浮点数"""
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def parse_subsidiary(org_name: str) -> tuple[str | None, str | None]:
    """从项目单位名称中提取子公司和城市信息"""
    if not org_name:
        return None, None

    cities = ['唐山', '承德', '张家口', '秦皇岛', '廊坊', '北京']
    found_city = None
    for c in cities:
        if c in org_name:
            found_city = c
            break
    return org_name, found_city


# ============================================================
# 主流水线
# ============================================================
def main():
    print("ECP2.0 完整数据采集流水线")
    print(f"目标: 国网冀北电力有限公司")
    print(f"开始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"下载URL: {ECP_DOWNLOAD}?noticeId={{noticeId}}&noticeDetId=")

    # 初始化数据库
    conn = init_db()
    cursor = conn.cursor()
    print("数据库已初始化")

    # 记录开始
    cursor.execute("""
        INSERT INTO crawl_log (task_name, task_type, status)
        VALUES ('pipeline_full', 'full_pipeline', 'running')
    """)
    conn.commit()

    # Phase 1+2
    material_notices = phase1_collect_list(conn)
    if not material_notices:
        # 从数据库读取已有的物资公告
        cursor.execute("""
            SELECT notice_id, title, notice_publish_time, code
            FROM bid_notices WHERE category='material'
            ORDER BY notice_publish_time DESC
        """)
        material_notices = [(r[0], r[1], r[2], r[3]) for r in cursor]
        print(f"从数据库读取 {len(material_notices)} 条物资公告")

    # Phase 3+4: 下载+解析+入库 (限制数量避免过载)
    MAX_DOWNLOAD = 10  # 验证阶段限制10个
    success_count = 0
    total_items = 0
    all_cities = Counter()
    all_materials = Counter()
    all_units = Counter()

    print(f"\n{'='*70}")
    print(f"Phase 3+4: 下载+解析货物清单 (验证 {min(MAX_DOWNLOAD, len(material_notices))} 条)")
    print(f"{'='*70}")

    for i, (notice_id, title, pub_time, code) in enumerate(material_notices[:MAX_DOWNLOAD]):
        print(f"\n[{i+1}/{min(MAX_DOWNLOAD, len(material_notices))}] {pub_time} {title[:70]}...")

        # 下载
        zip_path = download_zip(notice_id)
        if not zip_path:
            print(f"  -> 下载失败或文件为空")
            continue

        size_kb = os.path.getsize(zip_path) / 1024
        if size_kb < 100:  # 空ZIP (变更公告等)
            print(f"  -> 空ZIP ({size_kb:.0f}KB), 跳过")
            continue
        print(f"  -> 下载成功 ({size_kb:.0f}KB)")

        # 解压
        xlsx_path, doc_path = extract_zip(zip_path)
        if not xlsx_path:
            print(f"  -> 未找到货物清单XLSX")
            continue

        # 获取文件名 (处理Win32 unzip的乱码)
        raw_name = os.path.basename(xlsx_path)
        safe_name = raw_name.encode('utf-8', errors='replace').decode('utf-8', errors='replace')
        print(f"  -> XLSX: {safe_name} ({os.path.getsize(xlsx_path)/1024:.0f}KB)")

        # 解析
        items = parse_xlsx(xlsx_path)
        print(f"  -> 解析: {len(items)} 条物资, {len(set(i['sub_bid_name'] for i in items))} 个分标")

        # 入库 bid_items
        for item in items:
            # 存储到 bid_items，附加字段放入 remark
            extra_info = ""
            if item.get('delivery_method'):
                extra_info += f"交货方式:{item['delivery_method']}; "
            if item.get('voltage_level'):
                extra_info += f"电压等级:{item['voltage_level']}; "
            if item.get('project_name'):
                extra_info += f"项目:{item['project_name']}; "

            cursor.execute("""
                INSERT INTO bid_items
                    (notice_id, sub_bid_code, sub_bid_name, package_no,
                     material_name, material_desc, demand_quantity, unit,
                     project_org_name, delivery_place, remark,
                     source, source_file)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                notice_id, item['sub_bid_code'], item['sub_bid_name'],
                item['package_no'],
                item['material_name'], item['material_desc'],
                item['demand_quantity'], item['unit'],
                item['project_org_name'], item['delivery_place'],
                extra_info.strip() or None,
                item['source'], item['source_file'],
            ))

            # 统计
            if item['material_name']:
                # 取物资简称 (不含规格参数)
                short_name = item['material_name'].split(',')[0].split('(')[0].strip()
                all_materials[short_name] += 1
            if item['unit']:
                all_units[item['unit']] += 1
            if item['project_org_name']:
                _, city = parse_subsidiary(item['project_org_name'])
                if city:
                    all_cities[city] += 1

        conn.commit()

        # 入库 org_units (子公司)
        org_names = set()
        for item in items:
            if item['project_org_name']:
                org_names.add(item['project_org_name'])
            if item['demand_org_name']:
                org_names.add(item['demand_org_name'])

        for org_name in org_names:
            _, city = parse_subsidiary(org_name)
            cursor.execute("""
                INSERT OR IGNORE INTO org_units
                    (parent_org_id, parent_org_name, org_name, city, province,
                     first_seen_notice_id, first_seen_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                TARGET_ORG_ID, "国网冀北电力有限公司", org_name, city,
                "河北省" if city and city != "北京" else "北京市",
                notice_id, pub_time,
            ))

        conn.commit()

        # 标记详情已采集
        cursor.execute("""
            UPDATE bid_notices SET detail_fetched=1, detail_fetched_at=datetime('now')
            WHERE notice_id=?
        """, (notice_id,))
        conn.commit()

        success_count += 1
        total_items += len(items)

    # ============================================================
    # 汇总报告
    # ============================================================
    print(f"\n{'='*70}")
    print("采集汇总")
    print(f"{'='*70}")

    cursor.execute("SELECT COUNT(*) FROM bid_notices WHERE category='material'")
    mat_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM bid_items")
    item_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT notice_id) FROM bid_items")
    notice_with_items = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM org_units")
    org_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM bid_notices WHERE detail_fetched=1")
    detail_fetched = cursor.fetchone()[0]

    print(f"""
    +------------------------------------------------------+
    |  最终数据库状态                                       |
    +------------------------------------------------------+
    |  物资公告总数:        {mat_count:>4} 条
    |  已下载详情:          {detail_fetched:>4} 条 (本次下载: {success_count})
    |  物资明细条目:        {item_count:>5} 条
    |  有明细的公告:        {notice_with_items:>4} 条
    |  子公司/项目单位:     {org_count:>4} 个
    +------------------------------------------------------+
    """)

    print(f"物资名称分布 (Top 15):")
    for name, cnt in all_materials.most_common(15):
        print(f"  {name}: {cnt} 条")

    print(f"\n计量单位分布:")
    for unit, cnt in all_units.most_common():
        print(f"  {unit}: {cnt} 条")

    print(f"\n城市分布 (子公司维度):")
    for city, cnt in all_cities.most_common():
        print(f"  {city}: {cnt} 条")

    # 列出所有子公司
    cursor.execute("SELECT org_name, city FROM org_units ORDER BY city, org_name")
    print(f"\n已发现的子公司/项目单位:")
    for row in cursor:
        print(f"  [{row[1] or '未知'}] {row[0][:60]}")

    # 导出JSON样本
    cursor.execute("""
        SELECT notice_id, sub_bid_name, package_no, material_name,
               material_desc, demand_quantity, unit, project_org_name,
               delivery_place
        FROM bid_items LIMIT 50
    """)
    sample = []
    for row in cursor:
        sample.append({
            'notice_id': str(row[0]),
            'sub_bid_name': row[1],
            'package_no': row[2],
            'material_name': row[3],
            'material_desc': row[4],
            'demand_quantity': row[5],
            'unit': row[6],
            'project_org_name': row[7],
            'delivery_place': row[8],
        })

    sample_path = os.path.join(
        os.path.dirname(__file__), "..", "data", "samples", "bid_items_sample.json"
    )
    with open(sample_path, 'w', encoding='utf-8') as f:
        json.dump(sample, f, ensure_ascii=False, indent=2)
    print(f"\n样本导出: {sample_path} ({len(sample)} 条)")

    # 更新日志
    cursor.execute("""
        UPDATE crawl_log SET status='completed', records_count=?,
        completed_at=datetime('now')
        WHERE task_name='pipeline_full' AND status='running'
    """, (item_count,))
    conn.commit()

    conn.close()

    print(f"\n完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据库: {os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'ecp_data.db'))}")


if __name__ == "__main__":
    main()

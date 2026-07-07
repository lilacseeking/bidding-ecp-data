"""
从ECP数据库提取真实采购数据，生成 vmd-catboost 模型输入文件。

输出: vmd-catboost/inputs/data/data.xlsx
- 5个Sheet: Top5采购频率最高物资
- 列: 日期, 需求量, 项目数量(同源,论文披露),
       transformer_bids, monthly_bid_count, uhv_bids,
       has_batch, digital_bids, org_id
"""
import os, sys, io, sqlite3, numpy as np
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BIDDING_ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BIDDING_ROOT, "data", "ecp_data.db")
VMD_DATA_DIR = r"C:\Users\董文涛\PycharmProjects\vmd-catboost\inputs"
OUTPUT = os.path.join(VMD_DATA_DIR, "data.xlsx")

START = "201911"
END = "202607"
SGCC_ORG_ID = "2019040100044796"


def all_months(start, end):
    months = []
    y, m = int(start[:4]), int(start[4:])
    y_e, m_e = int(end[:4]), int(end[4:])
    while (y < y_e) or (y == y_e and m <= m_e):
        months.append("{:04d}{:02d}".format(y, m))
        m += 1
        if m > 12: m = 1; y += 1
    return months


def get_top5_materials(conn):
    """获取采购频率最高的5种物资"""
    c = conn.cursor()
    c.execute("""
        SELECT material_name, COUNT(DISTINCT demand_month) as months,
               SUM(demand_quantity) as total_qty
        FROM material_demand_total
        WHERE org_id=?
        GROUP BY material_name
        ORDER BY months DESC, total_qty DESC
        LIMIT 20
    """, (SGCC_ORG_ID,))
    top = c.fetchall()
    print("Top 20 materials by procurement frequency:")
    for i, (name, m, qty) in enumerate(top):
        mark = " ← SELECTED" if i < 5 else ""
        print(f"  #{i+1}: {name} ({m} months, {qty:.0f} qty){mark}")
    return [r[0] for r in top[:5]]


def load_batch_features():
    """加载批次事件特征（从 fetch_batch_features.py 生成的CSV，或直接从DB提取）"""
    import re
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT MIN(notice_publish_time), MAX(notice_publish_time) FROM bid_notices")
    min_pt, max_pt = c.fetchone()
    months = all_months(min_pt[:4]+min_pt[5:7], max_pt[:4]+max_pt[5:7])

    c.execute("""SELECT notice_id, title, notice_publish_time
        FROM bid_notices WHERE category='material' AND doctype='doci-bid'""")
    bids = c.fetchall()
    conn.close()

    from collections import defaultdict
    monthly = defaultdict(lambda: {'transformer': 0, 'uhv': 0, 'digital': 0, 'meter': 0, 'power': 0, 'total': 0})
    for nid, title, pt in bids:
        ym = pt[:4] + pt[5:7]
        monthly[ym]['total'] += 1
        if any(k in title for k in ['输变电', '变电设备']):
            monthly[ym]['transformer'] += 1
        if '特高压' in title: monthly[ym]['uhv'] += 1
        if '数字化' in title: monthly[ym]['digital'] += 1
        if '电源' in title: monthly[ym]['power'] += 1

    result = {}
    for ym in months:
        d = monthly.get(ym, defaultdict(int))
        result[ym] = {
            'transformer_bids': d['transformer'],
            'monthly_bid_count': d['total'],
            'uhv_bids': d['uhv'],
            'digital_bids': d['digital'],
            'has_batch': 1 if d['total'] > 0 else 0,
        }
    return result


def main():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Determine target materials
    target_mats = get_top5_materials(conn)
    print(f"\n目标物资 (Top5): {target_mats}")

    # Load material demand data
    c.execute("""
        SELECT material_name, demand_month,
               SUM(demand_quantity) as total_qty,
               MAX(notice_count) as notice_cnt
        FROM material_demand_total
        WHERE org_id=? AND material_name IN ({})
        GROUP BY material_name, demand_month
        ORDER BY material_name, demand_month
    """.format(",".join(["?"]*len(target_mats))), [SGCC_ORG_ID] + target_mats)
    rows = c.fetchall()
    conn.close()

    # Load batch features
    batch_feats = load_batch_features()

    months = all_months(START, END)
    print(f"数据范围: {START} ~ {END}, 共 {len(months)} 个月")

    # Organize by material
    mat_data = {mat: {} for mat in target_mats}
    for name, dm, qty, ncnt in rows:
        if name in mat_data:
            mat_data[name][dm] = (qty or 0, ncnt or 0)

    # Generate Excel
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADERS = [
        "日期", "需求量", "项目数量",
        "transformer_bids", "monthly_bid_count", "uhv_bids",
        "has_batch", "digital_bids",
    ]

    for mat in target_mats:
        ws = wb.create_sheet(mat)
        for ci, h in enumerate(HEADERS, 1):
            ws.cell(row=1, column=ci, value=h)

        zero_months = 0
        for ri, dm in enumerate(months, 2):
            qty, ncnt = mat_data[mat].get(dm, (0, 0))
            if qty == 0: zero_months += 1

            y, m = int(dm[:4]), int(dm[4:])
            bf = batch_feats.get(dm, {})

            ws.cell(row=ri, column=1, value=f"{y}-{m:02d}-01")
            ws.cell(row=ri, column=2, value=qty)
            ws.cell(row=ri, column=3, value=ncnt)
            ws.cell(row=ri, column=4, value=bf.get('transformer_bids', 0))
            ws.cell(row=ri, column=5, value=bf.get('monthly_bid_count', 0))
            ws.cell(row=ri, column=6, value=bf.get('uhv_bids', 0))
            ws.cell(row=ri, column=7, value=bf.get('has_batch', 0))
            ws.cell(row=ri, column=8, value=bf.get('digital_bids', 0))

        print(f"  [{mat}] {len(months)}月, 零值月={zero_months}, 非零月={len(months)-zero_months}")

    os.makedirs(VMD_DATA_DIR, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n输出: {OUTPUT}")


if __name__ == "__main__":
    main()
